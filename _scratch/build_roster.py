"""Build references/vc-roster.xlsx from the three CB Insights top-50 PDFs.

Two sheets:
  - All 150  (everything, with my best-effort sector tag)
  - Herb roster (filtered to climate / food / chem / industry-AI / deeptech)

The Herb-roster sheet is what gets consumed by Phase 2 / Source 2 of the
search playbook. NOT_RELEVANT funds (pure consumer / pure B2B SaaS) are
excluded. UNKNOWN entries kept in the All-150 sheet for review.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# (rank, name, country, sectors, fit, portfolio_url, note)
# fit values:
#   YES    — primary thesis includes climate / food / chem / industrial / industry-AI
#   MAYBE  — generalist deeptech, includes some relevant deals but not core thesis
#   NO     — consumer / pure B2B SaaS / fintech only — herb should skip
#   ?      — name not legible from PDF, needs user fill-in

# Bucket key:
#   CLIMATE  CCUS, energy, decarb
#   FOOD     food tech, agri, alt-protein, fermentation
#   CHEM     bio-based chemicals, advanced materials
#   IND      industrial / hardware / robotics
#   IND-AI   AI for industrial / process / Bühler-relevant
#   DEEPTECH generalist deep tech (some relevant exposure, hard to predict)

GERMANY = [
    (1, "High-Tech Gründerfonds (HTGF)", "DE", "DEEPTECH+IND", "YES", "https://www.htgf.de/en/portfolio/", "deep tech generalist incl. industrial bio + chem"),
    (2, "Bayern Kapital", "DE", "DEEPTECH", "MAYBE", "https://www.bayernkapital.de/en/portfolio", "Bavarian regional fund-of-funds, broad"),
    (3, "Global Founders Capital (GFC)", "DE", "GENERALIST", "NO", "https://www.globalfounderscapital.com/", "consumer + SaaS focus; Rocket Internet vehicle"),
    (4, "HV Capital", "DE", "GENERALIST", "MAYBE", "https://www.hvcapital.com/companies", "broad; occasional industrial"),
    (5, "La Famiglia", "DE", "B2B SAAS", "NO", "https://www.lafamiglia.vc/", "B2B SaaS focused"),
    (6, "Cherry Ventures", "DE", "CONSUMER+SAAS", "NO", "https://www.cherry.vc/", "consumer + SaaS"),
    (7, "Project A", "DE", "GENERALIST", "MAYBE", "https://www.project-a.com/portfolio/", "B2B + some industrial"),
    (8, "Picus Capital", "DE", "GENERALIST", "MAYBE", "https://www.picuscap.com/portfolio", "broad seed"),
    (9, "DTCP", "DE", "ENTERPRISE SAAS", "NO", "https://www.dtcp.capital/", "growth/SaaS focus"),
    (10, "Vsquared Ventures", "DE", "DEEPTECH+IND", "YES", "https://vsquared.vc/portfolio/", "deep tech, defense, industrial AI"),
    (11, "Atlantic Labs", "DE", "GENERALIST", "MAYBE", "https://atlanticlabs.com/", "B2B + some climate"),
    (12, "468 Capital", "DE", "DEEPTECH+CLIMATE", "YES", "https://www.468.capital/portfolio", "deep tech + climate explicit"),
    (13, "Redstone", "DE", "DEEPTECH", "MAYBE", "https://redstone.vc/portfolio/", "manages multiple thesis funds incl. climate"),
    (14, "IBB Ventures", "DE", "GENERALIST", "MAYBE", "https://www.ibbventures.de/en/portfolio", "Berlin regional"),
    (15, "b2venture", "DE", "GENERALIST", "MAYBE", "https://www.b2venture.vc/portfolio", "broad; some industrial"),
    (16, "[?] (UK-flag-pattern logo)", "DE", "UNKNOWN", "?", "", "user to identify — possibly a sector-specific or angel collective"),
    (17, "[?] (orange-dot logo)", "DE", "UNKNOWN", "?", "", "possibly Speedinvest? user to identify"),
    (18, "Foundamental", "DE", "IND", "YES", "https://www.foundamental.com/portfolio", "construction + industrial tech — strong fit"),
    (19, "Capnamic", "DE", "GENERALIST", "MAYBE", "https://www.capnamic.com/portfolio", "broad early-stage"),
    (20, "Leaps by Bayer", "DE", "FOOD+BIO", "YES", "https://leaps.bayer.com/portfolio/", "food, biotech, agri — direct LP fit"),
    (21, "Target Global", "DE", "GENERALIST", "MAYBE", "https://www.targetglobal.vc/portfolio/", "broad"),
    (22, "10x Founders", "DE", "GENERALIST", "MAYBE", "https://www.10xfounders.com/portfolio", "broad"),
    (23, "Robert Bosch Venture Capital (RBVC) / Bosch Ventures", "DE", "IND+IND-AI", "YES", "https://www.bosch-ventures.com/portfolio/", "industrial CVC — direct fit"),
    (24, "UVC Partners", "DE", "IND+DEEPTECH", "YES", "https://www.uvcpartners.com/portfolio", "deep tech industrial"),
    (25, "Acton Capital", "DE", "CONSUMER", "NO", "https://www.actoncapital.com/portfolio/", "consumer + commerce"),
    (26, "[?]", "DE", "UNKNOWN", "?", "", "user to identify"),
    (27, "Heartfelt", "DE", "CONSUMER", "NO", "https://www.heartfelt.vc/", "consumer"),
    (28, "TGFS (Technologiegründerfonds Sachsen)", "DE", "DEEPTECH", "MAYBE", "https://www.tgfs.de/portfolio/", "Saxony regional"),
    (29, "APX", "DE", "GENERALIST", "NO", "https://apx.vc/companies", "early-stage accelerator-style, broad"),
    (30, "FoodLabs", "DE", "FOOD+CLIMATE", "YES", "https://www.foodlabs.com/portfolio", "food + climate — direct fit"),
    (31, "BlueYard Capital", "DE", "DEEPTECH", "MAYBE", "https://blueyard.com/portfolio", "frontier deeptech"),
    (32, "eCAPITAL", "DE", "DEEPTECH+CLIMATE", "YES", "https://www.ecapital.de/en/portfolio/", "industrial tech, cleantech, software"),
    (33, "[?]", "DE", "UNKNOWN", "?", "", "user to identify"),
    (34, "BASF Venture Capital", "DE", "CHEM", "YES", "https://www.basf-vc.com/portfolio", "chemicals CVC — direct LP-adjacent"),
    (35, "Visionaries Club", "DE", "B2B SAAS", "NO", "https://www.visionariesclub.vc/", "pure B2B SaaS"),
    (36, "Hitachi Ventures", "DE", "IND+IND-AI", "YES", "https://www.hitachiventures.com/portfolio", "industrial CVC — direct fit"),
    (37, "Coparion", "DE", "DEEPTECH", "MAYBE", "https://www.coparion.vc/en/portfolio/", "deep tech generalist"),
    (38, "Fly Ventures", "DE", "DEEPTECH+B2B", "MAYBE", "https://www.fly.vc/portfolio", "B2B + AI; some industrial"),
    (39, "iCAP / Investmentbank Berlin (?)", "DE", "GENERALIST", "MAYBE", "", "regional — user to confirm"),
    (40, "SquareOne", "DE", "B2B SAAS", "NO", "https://www.squareone.vc/portfolio", "B2B SaaS"),
    (41, "CommerzVentures", "DE", "FINTECH", "NO", "https://commerzventures.com/portfolio/", "fintech"),
    (42, "NAP / Next Attractive Players (?)", "DE", "UNKNOWN", "?", "", "user to identify"),
    (43, "[?] (small v-logo)", "DE", "UNKNOWN", "?", "", "user to identify"),
    (44, "Vorwerk Ventures", "DE", "CONSUMER", "NO", "https://www.vorwerk-ventures.com/portfolio", "consumer/D2C"),
    (45, "EnBW New Ventures", "DE", "CLIMATE+ENERGY", "YES", "https://www.enbw-newventures.com/portfolio", "energy CVC — direct fit"),
    (45, "Boehringer Ingelheim Venture Fund (BIVF)", "DE", "BIO+FOOD", "YES", "https://www.boehringer-ingelheim.com/.../venture-fund", "biotech / food adjacent"),
    (47, "BonVenture", "DE", "IMPACT", "MAYBE", "https://www.bonventure.de/en/portfolio/", "impact-first; some climate/social"),
    (48, "D11Z Ventures", "DE", "GENERALIST", "MAYBE", "", "Stuttgart-area family office vehicle"),
    (49, "seed+speed", "DE", "GENERALIST", "MAYBE", "https://www.seedandspeed.com/portfolio/", "broad early-stage"),
    (50, "AENU", "DE", "CLIMATE", "YES", "https://www.aenu.com/portfolio/", "explicit climate fund — direct fit"),
]

UK = [
    (1, "Index Ventures", "UK", "GENERALIST", "MAYBE", "https://www.indexventures.com/companies/", "broad; rare industrial"),
    (2, "Phoenix Court (LocalGlobe / Latitude)", "UK", "GENERALIST", "MAYBE", "https://phoenixcourt.com/portfolio", "broad seed→growth"),
    (3, "Balderton Capital", "UK", "GENERALIST", "MAYBE", "https://www.balderton.com/portfolio", "broad"),
    (4, "Octopus Ventures", "UK", "CLIMATE+HEALTH", "YES", "https://octopusventures.com/our-companies/", "explicit deeptech + climate sleeve"),
    (5, "Seedcamp", "UK", "GENERALIST", "MAYBE", "https://seedcamp.com/portfolio/", "broad pre-seed/seed"),
    (6, "Atomico", "UK", "GENERALIST+CLIMATE", "MAYBE", "https://atomico.com/companies", "broad; growing climate exposure"),
    (7, "Ascension", "UK", "IMPACT+CONSUMER", "NO", "https://www.ascension.vc/our-investments", "impact tilt but consumer-heavy"),
    (8, "Frontline Ventures", "UK", "B2B SAAS", "NO", "https://www.frontline.vc/", "pure B2B SaaS"),
    (9, "Notion Capital", "UK", "B2B SAAS", "NO", "https://www.notion.vc/portfolio/", "pure B2B SaaS"),
    (10, "MMC Ventures", "UK", "DEEPTECH+IND-AI", "YES", "https://mmc.vc/portfolio/", "AI-first incl. industrial AI"),
    (11, "Felix Capital", "UK", "CONSUMER", "NO", "https://www.felixcap.com/portfolio", "consumer/lifestyle"),
    (12, "Northzone", "UK/SE", "GENERALIST", "MAYBE", "https://northzone.com/portfolio/", "broad"),
    (13, "Concept Ventures", "UK", "GENERALIST", "MAYBE", "https://conceptventures.vc/companies", "pre-seed broad"),
    (14, "Molten Ventures", "UK", "DEEPTECH", "MAYBE", "https://www.moltenventures.com/portfolio/", "deeptech — some industrial"),
    (15, "Entrée Capital", "UK", "GENERALIST", "MAYBE", "https://entreecap.com/portfolio/", "broad"),
    (16, "Fuel Ventures", "UK", "GENERALIST", "NO", "https://www.fuel.ventures/portfolio", "broad SaaS"),
    (17, "Coefficient Capital (?) / 'Coefficient'", "UK", "UNKNOWN", "?", "", "user to identify"),
    (18, "TS!C / 'TSIC'", "UK", "UNKNOWN", "?", "", "user to identify"),
    (19, "DN Capital", "UK", "GENERALIST", "NO", "https://www.dncapital.com/portfolio", "consumer + B2B SaaS"),
    (20, "Anthemis (?)", "UK", "FINTECH", "NO", "https://www.anthemis.com/portfolio", "fintech"),
    (21, "Amadeus Capital Partners", "UK", "DEEPTECH+IND", "YES", "https://www.amadeuscapital.com/portfolio/", "deeptech, industrial, semiconductors"),
    (22, "Connect Ventures", "UK", "CONSUMER+SAAS", "NO", "https://www.connectventures.co/portfolio", "consumer/SaaS"),
    (23, "Episode 1", "UK", "B2B SAAS", "NO", "https://www.episode1.com/portfolio", "B2B SaaS"),
    (24, "Hoxton Ventures (Haatch?)", "UK", "GENERALIST", "MAYBE", "https://www.hoxton.ventures/portfolio", "broad"),
    (25, "Beringea", "UK", "GENERALIST", "MAYBE", "https://www.beringea.com/portfolio/", "broad VCT-style"),
    (26, "AlbionVC (?)", "UK", "DEEPTECH", "MAYBE", "https://www.albion.vc/portfolio/", "B2B + some deep tech"),
    (27, "White Star Capital", "UK", "GENERALIST", "MAYBE", "https://whitestarcapital.com/portfolio/", "broad"),
    (28, "[?]", "UK", "UNKNOWN", "?", "", "user to identify"),
    (29, "83North", "UK", "GENERALIST", "MAYBE", "https://www.83north.com/portfolio", "broad B2B"),
    (30, "Fabric Ventures", "UK", "WEB3", "NO", "https://www.fabric.vc/portfolio", "crypto/web3"),
    (31, "[?] (group logo)", "UK", "UNKNOWN", "?", "", "user to identify"),
    (32, "Kingsworth (?)", "UK", "UNKNOWN", "?", "", "user to identify"),
    (33, "Playmaker (?)", "UK", "UNKNOWN", "?", "", "user to identify"),
    (34, "Dawn Capital", "UK", "B2B SAAS", "NO", "https://www.dawncapital.com/portfolio/", "pure B2B SaaS"),
    (35, "Episode 1 (dup?)", "UK", "B2B SAAS", "NO", "", "duplicate of #23 if present"),
    (36, "IQ Capital", "UK", "DEEPTECH+IND-AI", "YES", "https://www.iqcapital.vc/portfolio/", "deep tech + AI for industry"),
    (37, "Stride.VC (?)", "UK", "CONSUMER", "NO", "", "consumer-tilted"),
    (38, "[?]", "UK", "UNKNOWN", "?", "", "user to identify"),
    (39, "Moonfire Ventures", "UK", "GENERALIST", "MAYBE", "https://www.moonfire.com/portfolio", "broad seed"),
    (40, "AlbionVC", "UK", "DEEPTECH", "MAYBE", "https://www.albion.vc/portfolio/", "B2B + some deep tech (if not at #26)"),
    (41, "Hoxton Ventures", "UK", "GENERALIST", "MAYBE", "https://www.hoxton.ventures/portfolio", ""),
    (42, "Notion (?)", "UK", "UNKNOWN", "?", "", ""),
    (43, "Lightrock", "UK", "CLIMATE+IMPACT", "YES", "https://lightrock.com/portfolio/", "explicit climate + impact — direct fit"),
    (44, "Plant Capital (?) / pi-shaped logo", "UK", "UNKNOWN", "?", "", "user to identify"),
    (45, "Medicxi", "UK", "BIO", "MAYBE", "https://www.medicxi.com/portfolio/", "biotech only"),
    (46, "Playfair Capital", "UK", "GENERALIST", "MAYBE", "https://www.playfair.vc/portfolio", "broad seed"),
    (47, "Techstart Ventures", "UK", "GENERALIST", "NO", "https://www.techstart.vc/portfolio", "Northern Ireland regional"),
    (48, "Venrex", "UK", "CONSUMER", "NO", "https://www.venrex.com/portfolio/", "consumer"),
    (49, "Mercia / Mercia Asset Management", "UK", "GENERALIST", "MAYBE", "https://www.mercia.co.uk/portfolio/", "broad UK regional"),
    (50, "Augmentum Fintech", "UK", "FINTECH", "NO", "https://www.augmentum.vc/portfolio/", "fintech only"),
]

FRANCE = [
    (1, "Partech", "FR", "GENERALIST", "MAYBE", "https://partechpartners.com/companies/", "broad"),
    (2, "Isai", "FR", "GENERALIST", "MAYBE", "https://www.isai.fr/portfolio/", "broad"),
    (3, "Frst (fst)", "FR", "GENERALIST", "MAYBE", "https://frst.com/portfolio", "seed"),
    (4, "Sofinnova Partners", "FR", "DEEPTECH+BIO+CHEM", "YES", "https://www.sofinnovapartners.com/portfolio", "deeptech + industrial bio + chem — direct fit"),
    (5, "Elaia", "FR", "DEEPTECH", "MAYBE", "https://www.elaia.com/portfolio", "deeptech — broad"),
    (6, "[?] (NV-style logo)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (7, "Seventure Partners", "FR", "FOOD+BIO+CHEM", "YES", "https://www.seventure.fr/portfolio/", "food + microbiome + bio — direct fit"),
    (8, "Go Capital", "FR", "DEEPTECH", "MAYBE", "https://www.gocapital.fr/portfolio/", "deeptech regional"),
    (9, "360 Capital", "FR/IT", "GENERALIST", "MAYBE", "https://www.360capital.com/portfolio", "broad"),
    (10, "Supernova Invest", "FR", "DEEPTECH+IND", "YES", "https://www.supernovainvest.com/portfolio/", "deep tech + industrial — strong fit"),
    (11, "Serena", "FR", "GENERALIST", "NO", "https://www.serena.vc/portfolio", "broad SaaS"),
    (12, "Breega", "FR", "GENERALIST", "MAYBE", "https://www.breega.com/portfolio", "broad"),
    (13, "Newfund", "FR", "GENERALIST", "NO", "https://www.newfund.fr/portfolio/", "broad"),
    (14, "[?] (Cure Forge / similar)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (15, "AVP (Axa Venture Partners)", "FR", "GENERALIST+FINTECH", "MAYBE", "https://www.axavp.com/portfolio/", "fintech + broad"),
    (16, "Daphni", "FR", "GENERALIST", "MAYBE", "https://www.daphni.com/portfolio", "broad"),
    (17, "Ventech", "FR", "GENERALIST", "MAYBE", "https://www.ventechvc.com/portfolio", "broad"),
    (18, "Eleven (Eleven Ventures?)", "FR", "UNKNOWN", "?", "", "user to identify — could be 'eleven' / 'elaia'"),
    (19, "Founders Future", "FR", "GENERALIST", "NO", "https://www.foundersfuture.com/portfolio", "broad B2B"),
    (20, "AFI Ventures", "FR", "UNKNOWN", "?", "", "user to identify"),
    (21, "Xange", "FR", "GENERALIST", "MAYBE", "https://www.xange.vc/portfolio", "broad"),
    (22, "IRIS Capital", "FR", "GENERALIST+IND", "MAYBE", "https://www.iriscapital.com/portfolio", "broad incl. some industrial"),
    (23, "Aglaé Ventures", "FR", "CONSUMER", "NO", "https://www.aglaeventures.com/portfolio", "consumer + luxury (LVMH-related)"),
    (24, "Ring Capital", "FR", "GENERALIST", "MAYBE", "https://www.ringcm.com/portfolio", "broad"),
    (25, "[?] (Karista? Korelya?)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (26, "Axeleo Capital", "FR", "GENERALIST+IND-AI", "MAYBE", "https://www.axeleocapital.com/portfolio", "broad — some industrial AI"),
    (27, "Karista", "FR", "DEEPTECH+BIO", "MAYBE", "https://www.karista.vc/portfolio/", "deeptech + biotech"),
    (28, "[?]", "FR", "UNKNOWN", "?", "", "user to identify"),
    (29, "Brighteye Ventures", "FR", "EDTECH", "NO", "https://www.brighteyevc.com/", "edtech-only"),
    (30, "Innovacom", "FR", "DEEPTECH+IND", "MAYBE", "https://www.innovacom.com/portfolio/", "deeptech industrial"),
    (31, "Aquiti Gestion", "FR", "GENERALIST", "MAYBE", "", "regional South-West France"),
    (32, "Ragtime", "FR", "B2B SAAS", "NO", "https://www.ragtime.vc/", "B2B SaaS"),
    (33, "Side Capital", "FR", "GENERALIST", "NO", "", "broad"),
    (34, "Techmind (?)", "FR", "DEEPTECH", "MAYBE", "", "user to confirm"),
    (35, "Quanta Capital (?)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (36, "Jeito Capital", "FR", "BIO", "MAYBE", "https://www.jeito.life/portfolio/", "biotech"),
    (37, "BreizhUp", "FR", "GENERALIST", "MAYBE", "", "Brittany regional"),
    (38, "Cap Horn", "FR", "GENERALIST+IND-AI", "MAYBE", "https://www.caphorn.fr/portfolio/", "B2B + some industrial AI"),
    (39, "Express Partners (?) / EXES", "FR", "UNKNOWN", "?", "", "user to identify"),
    (40, "Nox Alpha (?)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (41, "Sharpstone Capital", "FR", "GENERALIST", "MAYBE", "", "broad"),
    (42, "[?]", "FR", "UNKNOWN", "?", "", "user to identify"),
    (43, "Odyssée Venture", "FR", "GENERALIST", "MAYBE", "https://www.odyssee-venture.com/portfolio", "broad"),
    (44, "MAIF Avenir", "FR", "INSURTECH", "NO", "https://www.maifavenir.fr/portfolio", "insurance CVC"),
    (45, "Söderberg & Partners (?)", "FR/SE", "UNKNOWN", "?", "", "user to identify"),
    (46, "[?] (purple-circle logo)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (47, "iD4 Ventures (?)", "FR", "UNKNOWN", "?", "", "user to identify"),
    (47, "Singular", "FR", "DEEPTECH", "MAYBE", "https://www.singular.vc/portfolio", "deep tech + AI"),
    (50, "Citizen Capital", "FR", "IMPACT", "MAYBE", "https://www.citizencapital.fr/portfolio/", "impact"),
    (50, "Finovam", "FR", "GENERALIST", "MAYBE", "", "Normandy regional"),
]

ALL_ROWS = []
for src, country in [(GERMANY, "DE"), (UK, "UK"), (FRANCE, "FR")]:
    for row in src:
        rank, name, c, sectors, fit, url, note = row
        ALL_ROWS.append((country, rank, name, sectors, fit, url, note))


def main():
    out = Path.home() / "herb-cloud" / "_scratch" / "vc-roster.xlsx"
    wb = Workbook()
    BOLD = Font(bold=True)
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREY  = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    WRAP  = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: All 150
    ws = wb.active
    ws.title = "All 150"
    headers = ["Country", "Rank", "VC", "Sectors (best-effort)", "Fit", "Portfolio URL", "Note"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = BOLD; c.alignment = WRAP
    for i, (country, rank, name, sectors, fit, url, note) in enumerate(ALL_ROWS, 2):
        for j, v in enumerate([country, rank, name, sectors, fit, url, note], 1):
            ws.cell(row=i, column=j, value=v).alignment = WRAP
        fc = ws.cell(row=i, column=5)
        fc.fill = {"YES": GREEN, "MAYBE": AMBER, "NO": RED, "?": GREY}.get(fit, None) or GREEN
    widths = [8, 6, 36, 22, 8, 48, 60]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 28

    # Sheet 2: Herb roster (filtered to YES + MAYBE)
    ws2 = wb.create_sheet("Herb roster (filtered)")
    for j, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = BOLD; c.alignment = WRAP
    rownum = 2
    for country, rank, name, sectors, fit, url, note in ALL_ROWS:
        if fit not in ("YES", "MAYBE"):
            continue
        for j, v in enumerate([country, rank, name, sectors, fit, url, note], 1):
            ws2.cell(row=rownum, column=j, value=v).alignment = WRAP
        ws2.cell(row=rownum, column=5).fill = GREEN if fit == "YES" else AMBER
        rownum += 1
    for j, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.row_dimensions[1].height = 28

    wb.save(out)

    # Stats
    by_fit = {"YES": 0, "MAYBE": 0, "NO": 0, "?": 0}
    for r in ALL_ROWS:
        by_fit[r[4]] = by_fit.get(r[4], 0) + 1
    print(f"Saved {out}")
    print(f"Total entries: {len(ALL_ROWS)}")
    for k, v in by_fit.items():
        print(f"  {k}: {v}")
    print(f"Herb roster (YES + MAYBE): {by_fit['YES'] + by_fit['MAYBE']}")


if __name__ == "__main__":
    main()
